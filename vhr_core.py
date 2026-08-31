"""
VHR core - STBET UK Virtual race card scraper + ODDS-matrix Excel writer.

The STBET card is published progressively: at 04:30 only the next ~30 minutes
of races exist, and the rest of the 24-hour card fills in as the day goes on
(verified against results history - a 04:30-only scrape misses ~30% of the
races that actually run).  So everything here is built around *merging*: each
poll adds the races that have appeared since the last one into a per-day state
file, and the Excel sheet is rebuilt from that state.

State lives in state/<card-date>.json so a restart never loses collected races.
"""

import json
import os
import re
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.page import PageMargins
from datetime import datetime, date, time as dtime, timedelta
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DAYS_DIR  = os.path.join(BASE_DIR, "days")
STATE_DIR = os.path.join(BASE_DIR, "state")

# The card rolls over at 04:30 LK - a "race day" runs 04:30 today -> 04:30 tomorrow.
CARD_ROLLOVER = dtime(4, 30)

# Short prices are hidden (not deleted) in the sheet, same as the old workbook.
ODDS_HIDE = {
    "5/4", "6/4", "13/8", "7/4", "2/1", "9/4", "5/2", "11/4",
    "3/1", "7/2", "4/1", "9/2", "5/1", "11/2", "6/1", "13/2",
    "7/1", "15/2", "10/3",
}

TRACK_COLORS = {
    "Portman Park":      ("8B0000", "F5E6E6"),
    "Sprint Valley":     ("1A3A6B", "E6ECF5"),
    "Steepledowns":      ("1B5E20", "E6F5E6"),
    "Canterberry Hills": ("6A1B9A", "F0E6F5"),
    "Home Straight":     ("B34700", "F5EAE1"),
    "Sandy Lanes":       ("8B6914", "F5F0E1"),
}
FALLBACK_COLORS = [("00695C", "E0F2F1"), ("4E342E", "EFEBE9"), ("37474F", "ECEFF1")]
TRACK_ORDER = ["Portman Park", "Sprint Valley", "Steepledowns"]

# Side meetings with a handful of races are not part of the three main cards.
MIN_TRACK_EVENTS = 30

STBET_BASE = "https://www.stbet.com/stbetrest/services/online/"
HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36"),
    "Accept": "application/json",
    "Referer": "https://www.stbet.com/",
}


# -- card day -----------------------------------------------------------------

def card_day(now=None):
    """The race-card date for a moment in time. Before 04:30 we are still on
    yesterday's card, because the card runs 04:30 -> 04:30."""
    now = now or datetime.now()
    d = now.date()
    if now.time() < CARD_ROLLOVER:
        d -= timedelta(days=1)
    return d


def card_window(day):
    """(start, end) datetimes of the card that starts on `day` at 04:30."""
    start = datetime.combine(day, CARD_ROLLOVER)
    return start, start + timedelta(days=1)


# -- STBET API ----------------------------------------------------------------

def _date_param():
    """The site's own formatDate quirk puts minutes in the middle slot; the
    server accepts it, so send the same shape it does."""
    n = datetime.now()
    h = n.hour % 12 or 12
    return f"{n:%d}/{n:%M}/{n.year} {h}:{n:%M}:{n:%S}"


def _params(**extra):
    p = {"country": "VR", "date": _date_param(),
         "filter": "2", "inplay": "false", "sport": "HR"}
    p.update(extra)
    return p


def api_get(path, params=None, timeout=20):
    r = requests.get(STBET_BASE + path, headers=HEADERS, params=params, timeout=timeout)
    r.raise_for_status()
    data = r.json()
    if isinstance(data, dict) and "payload" in data:
        return data["payload"]
    return data


def clean_track_name(meeting_name):
    """'SprintValley (VR)' -> 'Sprint Valley'."""
    name = (meeting_name or "").replace("(VR)", "").strip()
    return {"SprintValley": "Sprint Valley"}.get(name, name)


def normalize_odds(sel):
    po = sel.get("priceOdds") or {}
    frac = (po.get("priceFraction") or "").strip()
    if not frac or po.get("id") == -1:
        return "N/A"
    if frac.lower() in ("1/1", "evs", "evens"):
        return "Evs"
    return frac


def odds_value(o):
    """Fractional odds -> decimal, so columns sort by real price (this is what
    put 80/1 after 100/1 in the old sheet: unknown prices were just appended)."""
    if (o or "").lower() == "evs":
        return 1.0
    m = re.fullmatch(r"\s*(\d+)\s*/\s*(\d+)\s*", o or "")
    if not m:
        return float("inf")
    a, b = int(m.group(1)), int(m.group(2))
    return a / b if b else float("inf")


def get_meetings():
    """{track_name: meeting_id} for every VR meeting currently on the site."""
    payload = api_get("meeting/activeMeetings", _params())
    out = {}
    for m in payload or []:
        name = clean_track_name(m.get("meetingName", ""))
        if name and m.get("id"):
            out[name] = m["id"]          # duplicates in the feed share one id
    return out


def _resolve_date(label, day):
    """'Aug 31' -> a real date near `day`. The feed gives no year, and around
    the 04:30 rollover the old card is still being served, so a label has to
    resolve to its own real date rather than 'today or tomorrow'."""
    for yr in (day.year, day.year + 1, day.year - 1):
        try:
            d = datetime.strptime(f"{label} {yr}", "%b %d %Y").date()
        except (ValueError, TypeError):
            continue
        if abs((d - day).days) <= 3:
            return d
    return day


def get_track_events(meeting_id, day):
    """[{id, date, time}] for one meeting, dates resolved to real ISO dates."""
    payload = api_get("meeting/allEventsForMeeting", _params(meetingId=str(meeting_id)))
    label_today = f"{day:%b} {day.day}"       # matches eventLocalDate, e.g. "Aug 31"
    events, seen = [], set()
    for group in payload or []:
        for ev in group.get("eventLightDTOs", []):
            eid = ev.get("id")
            if not eid or eid in seen:
                continue
            seen.add(eid)
            ev_date = _resolve_date(ev.get("eventLocalDate", label_today), day)
            events.append({"id": eid,
                           "date": ev_date.isoformat(),
                           "time": (ev.get("eventTime") or "")[:5]})
    return [e for e in events if e["time"]]


def normalize_fraction(frac):
    """'9/1' / '1/1' / '' -> display odds. Results carry a bare string, not the
    priceOdds object the card uses."""
    frac = (frac or "").strip()
    if not frac:
        return None
    if frac.lower() in ("1/1", "evs", "evens"):
        return "Evs"
    return frac


def fetch_event_result(event_id):
    """Winner and placed odds for a settled race, or None if it has not run."""
    try:
        res = api_get("meeting/eventResultsByEventId", {"eventId": str(event_id)})
    except Exception:
        return None
    if not res:
        return None
    r0 = res[0]
    pos = {p.get("resultPosition"): p for p in r0.get("eventPositions", [])}
    win = pos.get(1)
    if not win:
        return None
    return {
        "winner":      (win.get("selectionName") or "").strip(),
        "winner_odds": normalize_fraction(win.get("lastOdd")),
        "second_odds": normalize_fraction((pos.get(2) or {}).get("lastOdd")),
        "third_odds":  normalize_fraction((pos.get(3) or {}).get("lastOdd")),
        "runners":     r0.get("runners") or 0,
    }


def fetch_event_odds(event_id):
    """Odds of every runner in one race. selectionNumber 0 is the synthetic
    'Favourite' betting option, not a horse."""
    try:
        sels = api_get("meeting/eventSelections", {"eventId": str(event_id)})
    except Exception:
        return None
    odds = []
    for sel in sels or []:
        if sel.get("selectionNumber", 0) == 0:
            continue
        if (sel.get("name") or "").strip():
            odds.append(normalize_odds(sel))
    return odds


# -- state --------------------------------------------------------------------

def state_path(day):
    return os.path.join(STATE_DIR, f"{day.isoformat()}.json")


def load_state(day):
    p = state_path(day)
    if os.path.exists(p):
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"card_date": day.isoformat(), "tracks": {}}


def save_state(day, state):
    os.makedirs(STATE_DIR, exist_ok=True)
    tmp = state_path(day) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=1)
    os.replace(tmp, state_path(day))


def race_dt(rec):
    h, m = map(int, rec["time"].split(":"))
    return datetime.combine(date.fromisoformat(rec["date"]), dtime(h, m))


def card_status(day):
    """{track: {"total": n, "upcoming": n}} for the card the site is serving.

    'upcoming' - races that have not started yet - is what says whether the new
    24-hour card is up. STBET swaps the whole card in one go between 04:30 and
    05:00, and until it does, the finished card is still being served: plenty of
    races in total, almost none of them still to run. Counting totals is what
    made the old scheduler scrape the dead card at 04:30:02.
    """
    meetings = get_meetings()
    now = datetime.now()
    out = {}
    for track, mid in meetings.items():
        try:
            events = get_track_events(mid, day)
        except Exception:
            continue
        if len(events) < MIN_TRACK_EVENTS:
            continue                                   # side meeting
        out[track] = {"total": len(events),
                      "upcoming": sum(1 for e in events if race_dt(e) > now)}
    return out


def collect(day, state, log=print, skip_started=False):
    """One poll: merge every race currently on the card into `state`.

    Races already stored keep their odds - prices do not move once published
    (two scrapes 95 minutes apart matched on all 57 shared races). Races stored
    without odds are retried, because a race can appear on the card before its
    prices go up.

    Returns (n_new, n_filled).
    """
    meetings = get_meetings()
    if not meetings:
        raise RuntimeError("no VR meetings on STBET right now")

    start, end = card_window(day)
    cutoff = max(start, datetime.now()) if skip_started else start

    todo, n_new = [], 0
    for track, mid in meetings.items():
        try:
            events = get_track_events(mid, day)
        except Exception as e:
            log(f"    {track}: event list failed ({e})")
            continue
        if len(events) < MIN_TRACK_EVENTS:
            continue                                   # side meeting
        slot = state["tracks"].setdefault(track, {})
        for ev in events:
            key = str(ev["id"])
            when = race_dt(ev)
            if not (start <= when < end):              # outside this card
                continue
            rec = slot.get(key)
            if rec is None:
                if when < cutoff:
                    continue                           # already run before we started
                slot[key] = {"date": ev["date"], "time": ev["time"], "odds": []}
                n_new += 1
                todo.append((track, key))
            elif not rec["odds"]:
                todo.append((track, key))              # prices were not up yet

    n_filled = 0
    if todo:
        with ThreadPoolExecutor(max_workers=8) as ex:
            futures = {ex.submit(fetch_event_odds, k): (t, k) for t, k in todo}
            for fu in as_completed(futures):
                track, key = futures[fu]
                odds = fu.result()
                if odds:
                    state["tracks"][track][key]["odds"] = odds
                    n_filled += 1
    return n_new, n_filled


# -- Excel --------------------------------------------------------------------

def _track_style(track, fallback_idx=0):
    if track in TRACK_COLORS:
        return TRACK_COLORS[track]
    return FALLBACK_COLORS[fallback_idx % len(FALLBACK_COLORS)]


def _write_odds_sheet(wb, track, races, color, light):
    """races: [(label_time, [odds, ...])] already in running order."""
    ws = wb.create_sheet(f"{track} ODDS")

    found = {o for _, odds in races for o in odds if o not in ("N/A", "")}
    columns = sorted(found, key=lambda o: (odds_value(o), o))
    header = ["TIME"] + columns

    ws.column_dimensions["A"].width = 10
    for i in range(2, len(header) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10

    def header_row():
        ws.append(header)
        r = ws.max_row
        ws.row_dimensions[r].height = 28
        for c_idx in range(1, len(header) + 1):
            c = ws.cell(row=r, column=c_idx)
            c.font = Font(bold=True, size=18, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal="center", vertical="center")

    header_row()

    race_num = stripe = 0
    for race_time, odds in races:
        if race_num and race_num % 10 == 0:      # repeat the header every 10 races
            header_row()
            stripe = 0
        counts = defaultdict(int)
        for o in odds:
            counts[o] += 1
        ws.append([race_time] + [counts.get(o, 0) for o in columns])
        r = ws.max_row
        race_num += 1
        stripe += 1
        ws.row_dimensions[r].height = 28
        for c_idx in range(1, len(header) + 1):
            c = ws.cell(row=r, column=c_idx)
            c.font = Font(bold=True, size=18)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if stripe % 2 == 0:
                c.fill = PatternFill("solid", fgColor=light)

    for c_idx in range(2, len(header) + 1):
        if ws.cell(row=1, column=c_idx).value in ODDS_HIDE:
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].hidden = True

    ws.print_title_rows = "1:1"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    return len(races)


def build_workbook(day, state, out_path):
    """Write the ODDS-matrix workbook. Returns (path, {track: n_races}).

    If the file is open in Excel the save raises PermissionError; we fall back
    to a '... LIVE.xlsx' name so a poll never loses its data.
    """
    tracks = sorted(state["tracks"],
                    key=lambda n: (TRACK_ORDER.index(n) if n in TRACK_ORDER else 99, n))
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    counts, fb = {}, 0
    for track in tracks:
        recs = [r for r in state["tracks"][track].values() if r["odds"]]
        if not recs:
            continue
        recs.sort(key=race_dt)
        color, light = _track_style(track, fb)
        if track not in TRACK_COLORS:
            fb += 1
        counts[track] = _write_odds_sheet(
            wb, track, [(r["time"], r["odds"]) for r in recs], color, light)

    if not counts:
        return None, {}

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    try:
        wb.save(out_path)
        return out_path, counts
    except PermissionError:
        alt = out_path[:-5] + " LIVE.xlsx"
        wb.save(alt)
        return alt, counts


def workbook_path(day):
    return os.path.join(DAYS_DIR, f"VHR Racecards {day:%Y.%m.%d}.xlsx")
