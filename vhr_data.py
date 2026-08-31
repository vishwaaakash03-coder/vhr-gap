"""
VHR race store - one record per race, holding both halves of the analysis.

A gap analysis needs two facts about every race:

  * which odds were **offered** (from the racecard), and
  * which odds **won** (from the result).

They come from different endpoints. This module keeps them in one place, keyed
by track|date|time.

The only source used by default is `BET/state/*.json` - the cards this system
collected itself. The previous system's files can still be read, but only when
asked for explicitly (`python vhr_data.py --legacy`): its scheduler fired on the
previous day's finished card, so its workbooks are not trustworthy and must not
find their way back into the store by accident.

Store: BET/data/races.json
"""

import json
import os
import re
from datetime import date, datetime, timedelta

BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DATA_DIR  = os.path.join(BASE_DIR, "data")
STORE     = os.path.join(DATA_DIR, "races.json")
STATE_DIR = os.path.join(BASE_DIR, "state")

# The previous system, kept only as an import source.
LEGACY_DIR     = os.path.join(os.path.dirname(BASE_DIR), "race", "days")
LEGACY_RESULTS = os.path.join(LEGACY_DIR, "results_history.json")

TRACK_ORDER = ["Portman Park", "Sprint Valley", "Steepledowns"]


def race_key(track, day, time_str):
    return f"{track}|{day}|{time_str}"


def parse_key(key):
    track, day, time_str = key.split("|")
    return track, day, time_str


def sort_key(rec):
    return (rec["date"], rec["time"])


# -- store ---------------------------------------------------------------------

def load_store():
    if os.path.exists(STORE):
        try:
            with open(STORE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {"races": {}}


def save_store(store):
    os.makedirs(DATA_DIR, exist_ok=True)
    tmp = STORE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(store, f, separators=(",", ":"))
    os.replace(tmp, STORE)


def upsert(store, track, day, time_str, **fields):
    """Merge fields into one race record, never clearing what is already known."""
    key = race_key(track, day, time_str)
    rec = store["races"].get(key)
    if rec is None:
        rec = {"track": track, "date": day, "time": time_str}
        store["races"][key] = rec
    for k, v in fields.items():
        if v not in (None, "", []):
            rec[k] = v
    return rec


def races_by_track(store, track=None):
    """{track: [record, ...]} in running order."""
    out = {}
    for rec in store["races"].values():
        if track and rec["track"] != track:
            continue
        out.setdefault(rec["track"], []).append(rec)
    for recs in out.values():
        recs.sort(key=sort_key)
    return out


# -- import: new collector state ------------------------------------------------

def import_state(store, days=None):
    """Cards collected by vhr_collector - exact odds, with event ids.

    `days` limits the import to those card dates. The results collector passes
    the current and previous card day so that every race it settles also carries
    the odds that were on offer - without that, races collected from tomorrow on
    would have a winner but no card, and quietly drop out of the BY CHANCES
    analysis.
    """
    n = 0
    if not os.path.isdir(STATE_DIR):
        return 0
    wanted = {d.isoformat() + ".json" for d in days} if days else None
    for fn in sorted(os.listdir(STATE_DIR)):
        if not fn.endswith(".json"):
            continue
        if wanted and fn not in wanted:
            continue
        try:
            with open(os.path.join(STATE_DIR, fn), "r", encoding="utf-8") as f:
                state = json.load(f)
        except Exception:
            continue
        for track, slot in state.get("tracks", {}).items():
            for event_id, rec in slot.items():
                odds = [o for o in rec.get("odds", []) if o and o != "N/A"]
                if not odds:
                    continue
                upsert(store, track, rec["date"], rec["time"],
                       odds=sorted(set(odds)), runners=len(rec["odds"]),
                       event_id=event_id, odds_src="state")
                n += 1
    return n


# -- import: legacy ODDS-matrix workbooks --------------------------------------

def import_legacy_xlsx(store, only_missing=True):
    """Older cards. Times run past midnight, so a time going backwards means the
    date has rolled over - that is how the real date of each row is recovered."""
    try:
        import openpyxl
    except ImportError:
        return 0
    if not os.path.isdir(LEGACY_DIR):
        return 0

    n = 0
    for fn in sorted(os.listdir(LEGACY_DIR)):
        m = re.match(r"VHR Racecards (\d{4})\.(\d{2})\.(\d{2})", fn)
        if not m or not fn.endswith(".xlsx") or fn.startswith("~"):
            continue
        day0 = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        try:
            wb = openpyxl.load_workbook(os.path.join(LEGACY_DIR, fn), read_only=True)
        except Exception:
            continue
        for sheet in wb.sheetnames:
            if not sheet.endswith(" ODDS"):
                continue
            track = sheet[:-5]
            header, day, prev = None, day0, None
            try:
                rows = list(wb[sheet].iter_rows(values_only=True))
            except Exception as e:
                # A few of the older workbooks are corrupt (bad CRC). Skip the
                # sheet rather than losing every later file to one bad zip.
                print(f"    skipped {fn} [{sheet}]: {type(e).__name__}")
                continue
            for row in rows:
                if not row or not row[0]:
                    continue
                if row[0] == "TIME":
                    header = [c for c in row]
                    continue
                if header is None:
                    continue
                time_str = str(row[0])
                if prev and time_str < prev:        # wrapped past midnight
                    day = day + timedelta(days=1)
                prev = time_str
                odds = sorted({header[i] for i, v in enumerate(row)
                               if i and isinstance(v, int) and v > 0 and header[i]})
                if not odds:
                    continue
                key = race_key(track, day.isoformat(), time_str)
                if only_missing and store["races"].get(key, {}).get("odds"):
                    continue
                upsert(store, track, day.isoformat(), time_str,
                       odds=odds, odds_src="xlsx")
                n += 1
        wb.close()
    return n


# -- import: legacy results -----------------------------------------------------

def import_legacy_results(store):
    """The old dashboard's results history - winner odds per race."""
    if not os.path.exists(LEGACY_RESULTS):
        return 0
    try:
        with open(LEGACY_RESULTS, "r", encoding="utf-8") as f:
            races = json.load(f).get("races", [])
    except Exception:
        return 0
    n = 0
    for r in races:
        if not r.get("winner_odds") or r["winner_odds"] == "N/A":
            continue
        upsert(store, r["track"], r["date"], r["time"],
               winner_odds=r["winner_odds"], winner=r.get("winner"),
               second_odds=r.get("second_odds"), third_odds=r.get("third_odds"),
               event_id=str(r.get("event_id") or "") or None,
               runners=r.get("runners"))
        n += 1
    return n


def rebuild(verbose=True, legacy=False):
    """Re-import the collected cards into the store. Safe to re-run.

    `legacy` additionally pulls the old system's workbooks and results history.
    Off by default - that data came from a scraper that read the wrong card.
    """
    store = load_store()
    say = print if verbose else (lambda *a, **k: None)

    n = import_state(store)
    say(f"  cards   (this system)      {n:6d} races")

    if legacy:
        n = import_legacy_results(store)
        say(f"  results (legacy history)   {n:6d} races")
        n = import_legacy_xlsx(store)
        say(f"  cards   (legacy workbooks) {n:6d} races")

    save_store(store)
    return store


def coverage(store):
    """How much of the store can each half of the analysis actually use."""
    total = len(store["races"])
    with_odds = sum(1 for r in store["races"].values() if r.get("odds"))
    with_win = sum(1 for r in store["races"].values() if r.get("winner_odds"))
    both = sum(1 for r in store["races"].values()
               if r.get("odds") and r.get("winner_odds"))
    days = sorted({r["date"] for r in store["races"].values()})
    return {"total": total, "with_odds": with_odds,
            "with_winner": with_win, "both": both,
            "since": days[0] if days else None,
            "until": days[-1] if days else None,
            "days": len(days)}


if __name__ == "__main__":
    import sys
    legacy = "--legacy" in sys.argv

    if "--reset" in sys.argv:
        if os.path.exists(STORE):
            backup = STORE + f".backup-{datetime.now():%Y%m%d-%H%M%S}"
            os.replace(STORE, backup)
            print(f"Old store moved to {os.path.basename(backup)}")
        else:
            print("No store to reset.")

    print("Rebuilding race store"
          + (" (including the old system's data)" if legacy else "") + " ...")
    store = rebuild(legacy=legacy)
    c = coverage(store)
    print(f"\nStore: {STORE}")
    print(f"  {c['total']:6d} races known")
    print(f"  {c['with_winner']:6d} with a winning odds  (gap by races)")
    print(f"  {c['with_odds']:6d} with the card's odds  (gap by chances)")
    print(f"  {c['both']:6d} with both")
    if c["since"]:
        print(f"  {c['since']} -> {c['until']}  ({c['days']} days)")
